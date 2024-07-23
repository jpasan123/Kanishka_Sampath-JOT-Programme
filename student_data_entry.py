import pandas as pd
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkcalendar import DateEntry  # Ensure you have tkcalendar installed
from PIL import Image, ImageTk

class StudentDataApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Job Oriented Training Programme Students & Data Entry")

        # Set window background color
        self.root.configure(bg='#f0f0f0')

        # Create a canvas and a scrollbar
        self.canvas = tk.Canvas(root, bg='#f0f0f0')
        self.scrollbar = tk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg='#f0f0f0')

        # Pack the widgets
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Load and place logo
        self.logo_image = Image.open("kanishka_logo.png")
        self.logo_image = self.logo_image.resize((200, 100), Image.Resampling.LANCZOS)
        self.logo_photo = ImageTk.PhotoImage(self.logo_image)
        self.logo_label = tk.Label(self.scrollable_frame, image=self.logo_photo, bg='#f0f0f0')
        self.logo_label.grid(row=0, column=0, columnspan=2, pady=10)

        # Label and Entry widget configuration
        label_font = ("Arial", 12, "bold")
        entry_font = ("Arial", 12)
        entry_width = 30
        label_bg = '#f0f0f0'
        label_fg = '#333333'

        # Helper function to create labels and entries
        def create_label_entry(row, text):
            label = tk.Label(self.scrollable_frame, text=text, font=label_font, bg=label_bg, fg=label_fg)
            label.grid(row=row, column=0, sticky='e', padx=10, pady=5)
            entry = tk.Entry(self.scrollable_frame, font=entry_font, width=entry_width)
            entry.grid(row=row, column=1, pady=5)
            entry.bind("<Return>", self.focus_next_widget)
            return entry

        # Labels and Entry widgets for Student Info
        self.student_no_entry = create_label_entry(1, "Student No")
        self.student_name_entry = create_label_entry(2, "Student Name")
        self.program_entry = create_label_entry(3, "Program")
        self.jot_activity1_entry = create_label_entry(4, "JOT Activity 1")
        self.phicycle_activity1_entry = create_label_entry(5, "Phicycle Activity 1")
        self.lab_section1_entry = create_label_entry(6, "Lab Section 1")
        self.assignment1_entry = create_label_entry(7, "Assignment 1")
        self.assessment1_entry = create_label_entry(8, "Assessment 1")
        self.practical1_entry = create_label_entry(9, "Practical 1")
        self.module1_title_entry = create_label_entry(10, "Module 1 Title")
        self.module1_subject1_entry = create_label_entry(11, "Module 1 Subject 1")

        # Module 1 Hours and Minutes
        self.module1_hours_label = tk.Label(self.scrollable_frame, text="Module 1 Hours", font=label_font, bg=label_bg, fg=label_fg)
        self.module1_hours_label.grid(row=12, column=0, sticky='e', padx=10, pady=5)
        self.module1_hours_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=23, wrap=True, state="readonly", width=5, font=entry_font)
        self.module1_hours_spinbox.grid(row=12, column=1, sticky='w', pady=5)
        self.module1_hours_spinbox.bind("<Return>", self.focus_next_widget)
        self.module1_minutes_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=59, wrap=True, state="readonly", width=5, font=entry_font)
        self.module1_minutes_spinbox.grid(row=12, column=1, sticky='e', pady=5)
        self.module1_minutes_spinbox.bind("<Return>", self.focus_next_widget)

        self.module1_supervisor_entry = create_label_entry(13, "Module 1 Supervisor")
        self.module1_lecturer_entry = create_label_entry(14, "Module 1 Lecturer")
        self.module1_satisfaction_entry = create_label_entry(15, "Lecturer or Supervisor Satisfaction")
        self.module1_lecturer_no_entry = create_label_entry(16, "Module Lecturer No")
        self.module1_supervisor_no_entry = create_label_entry(17, "Module Supervisor No")
        self.module1_title_no_entry = create_label_entry(18, "Module Title No")
        self.module1_subject_no_entry = create_label_entry(19, "Module Subject No")

        # Subject Hours and Minutes
        self.subject_hours_label = tk.Label(self.scrollable_frame, text="Subject Hours", font=label_font, bg=label_bg, fg=label_fg)
        self.subject_hours_label.grid(row=20, column=0, sticky='e', padx=10, pady=5)
        self.subject_hours_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=23, wrap=True, state="readonly", width=5, font=entry_font)
        self.subject_hours_spinbox.grid(row=20, column=1, sticky='w', pady=5)
        self.subject_hours_spinbox.bind("<Return>", self.focus_next_widget)
        self.subject_minutes_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=59, wrap=True, state="readonly", width=5, font=entry_font)
        self.subject_minutes_spinbox.grid(row=20, column=1, sticky='e', pady=5)
        self.subject_minutes_spinbox.bind("<Return>", self.focus_next_widget)

        self.module_date_label = tk.Label(self.scrollable_frame, text="Module Date", font=label_font, bg=label_bg, fg=label_fg)
        self.module_date_label.grid(row=21, column=0, sticky='e', padx=10, pady=5)
        self.module_date_entry = DateEntry(self.scrollable_frame, font=entry_font)
        self.module_date_entry.grid(row=21, column=1, pady=5)
        self.module_date_entry.bind("<Return>", self.focus_next_widget)

        self.subject_date_label = tk.Label(self.scrollable_frame, text="Subject Date", font=label_font, bg=label_bg, fg=label_fg)
        self.subject_date_label.grid(row=22, column=0, sticky='e', padx=10, pady=5)
        self.subject_date_entry = DateEntry(self.scrollable_frame, font=entry_font)
        self.subject_date_entry.grid(row=22, column=1, pady=5)
        self.subject_date_entry.bind("<Return>", self.focus_next_widget)

        self.final_approval_entry = create_label_entry(23, "Final Approval Note")

        # Submit Button
        self.submit_button = tk.Button(self.scrollable_frame, text="Submit", font=("Arial", 14, "bold"), bg="#4CAF50", fg="white", command=self.submit_data)
        self.submit_button.grid(row=24, column=0, columnspan=2, pady=10)

    def focus_next_widget(self, event):
        event.widget.tk_focusNext().focus()
        return "break"

    def submit_data(self):
        student_no = self.student_no_entry.get()
        student_name = self.student_name_entry.get()
        program = self.program_entry.get()
        jot_activity1 = self.jot_activity1_entry.get()
        phicycle_activity1 = self.phicycle_activity1_entry.get()
        lab_section1 = self.lab_section1_entry.get()
        assignment1 = self.assignment1_entry.get()
        assessment1 = self.assessment1_entry.get()
        practical1 = self.practical1_entry.get()
        module1_title = self.module1_title_entry.get()
        module1_subject1 = self.module1_subject1_entry.get()
        
        # Get values from Spinbox widgets
        module1_hours = f"{self.module1_hours_spinbox.get()}:{self.module1_minutes_spinbox.get()}"
        subject_hours = f"{self.subject_hours_spinbox.get()}:{self.subject_minutes_spinbox.get()}"
        
        module1_supervisor = self.module1_supervisor_entry.get()
        module1_lecturer = self.module1_lecturer_entry.get()
        module1_satisfaction = self.module1_satisfaction_entry.get()
        module1_lecturer_no = self.module1_lecturer_no_entry.get()
        module1_supervisor_no = self.module1_supervisor_no_entry.get()
        module1_title_no = self.module1_title_no_entry.get()
        module1_subject_no = self.module1_subject_no_entry.get()
        
        module_date = self.module_date_entry.get_date()
        subject_date = self.subject_date_entry.get_date()
        final_approval = self.final_approval_entry.get()

        # Load the existing Excel file or create a new one
        try:
            df = pd.read_excel('Student & Lecturer Data.xlsx')
        except FileNotFoundError:
            df = pd.DataFrame(columns=['Student No', 'Student Name', 'Program', 'JOT Activity 1', 
                                       'Phicycle Activity 1', 'Lab Section 1', 'Assignment 1', 
                                       'Assessment 1', 'Practical 1', 'Module 1 Title', 'Module 1 Subject 1', 
                                       'Module 1 Hours', 'Module 1 Supervisor', 'Module 1 Lecturer', 
                                       'Module 1 Satisfaction', 'Module Lecturer No', 'Module Supervisor No', 
                                       'Module Title No', 'Module Subject No', 'Subject Hours', 
                                       'Module Date', 'Subject Date', 'Final Approval Note'])

        new_data = pd.DataFrame([{
            'Student No': student_no,
            'Student Name': student_name,
            'Program': program,
            'JOT Activity 1': jot_activity1,
            'Phicycle Activity 1': phicycle_activity1,
            'Lab Section 1': lab_section1,
            'Assignment 1': assignment1,
            'Assessment 1': assessment1,
            'Practical 1': practical1,
            'Module 1 Title': module1_title,
            'Module 1 Subject 1': module1_subject1,
            'Module 1 Hours': module1_hours,
            'Module 1 Supervisor': module1_supervisor,
            'Module 1 Lecturer': module1_lecturer,
            'Module 1 Satisfaction': module1_satisfaction,
            'Module Lecturer No': module1_lecturer_no,
            'Module Supervisor No': module1_supervisor_no,
            'Module Title No': module1_title_no,
            'Module Subject No': module1_subject_no,
            'Subject Hours': subject_hours,
            'Module Date': module_date,
            'Subject Date': subject_date,
            'Final Approval Note': final_approval
        }])

        # Concatenate the new data with the existing DataFrame
        df = pd.concat([df, new_data], ignore_index=True)

        # Save to Excel file
        df.to_excel('Student & Lecturer Data.xlsx', index=False)

        # Highlight the new data in Excel
        self.highlight_new_data('Student & Lecturer Data.xlsx', len(df))

        messagebox.showinfo("Success", "Data submitted successfully!")
        self.clear_form()

    def highlight_new_data(self, file_name, row_number):
        workbook = load_workbook(file_name)
        sheet = workbook.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row_number + 1, column=col).fill = yellow_fill
        workbook.save(file_name)

    def clear_form(self):
        self.student_no_entry.delete(0, tk.END)
        self.student_name_entry.delete(0, tk.END)
        self.program_entry.delete(0, tk.END)
        self.jot_activity1_entry.delete(0, tk.END)
        self.phicycle_activity1_entry.delete(0, tk.END)
        self.lab_section1_entry.delete(0, tk.END)
        self.assignment1_entry.delete(0, tk.END)
        self.assessment1_entry.delete(0, tk.END)
        self.practical1_entry.delete(0, tk.END)
        self.module1_title_entry.delete(0, tk.END)
        self.module1_subject1_entry.delete(0, tk.END)
        self.module1_hours_spinbox.delete(0, tk.END)
        self.module1_minutes_spinbox.delete(0, tk.END)
        self.module1_supervisor_entry.delete(0, tk.END)
        self.module1_lecturer_entry.delete(0, tk.END)
        self.module1_satisfaction_entry.delete(0, tk.END)
        self.module1_lecturer_no_entry.delete(0, tk.END)
        self.module1_supervisor_no_entry.delete(0, tk.END)
        self.module1_title_no_entry.delete(0, tk.END)
        self.module1_subject_no_entry.delete(0, tk.END)
        self.subject_hours_spinbox.delete(0, tk.END)
        self.subject_minutes_spinbox.delete(0, tk.END)
        self.module_date_entry.set_date('')
        self.subject_date_entry.set_date('')
        self.final_approval_entry.delete(0, tk.END)

if __name__ == "__main__":
    root = tk.Tk()
    app = StudentDataApp(root)
    root.mainloop()
